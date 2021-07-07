import csv
import openpyxl as excel
import os
import datetime
def CSVWriter(parent,Data_ROW):
    #csvで実行する場合のメモ
    ID = parent.St_ID.get()
    ToDay = datetime.datetime.now().strftime("%Y-%m-%d")
    outputfile = open(os.getcwd()+"\OutPut\APPLYLIST_"+ID+"_"+ToDay+".csv","w",newline="")
    fieldnames = ["ID","名前","開始日","終了日","申請種別","申請理由"]
    Dict_Writer = csv.DictWriter(outputfile,fieldnames)
    Dict_Writer.writeheader()
    for data in Data_ROW:
        Dict_Writer.writerow({"ID":data["ID"],"名前":data["NAME"],"開始日":data["StartDate"],"終了日":data["EndDate"],"申請種別":data["Title"],"申請理由":data["Reason"]})
def MultiExcelWriter(Data_ROWS):
    """与えられた社員一覧を基にデータを抜き出しExcel出力を行う"""
    ToDay = datetime.datetime.now().strftime("%Y-%m-%d")
    OUTPUTFILENAME = os.getcwd()+"\OutPut\APPLYLISTS_"+ToDay+".xlsx"
    
    outputwb = excel.Workbook()
    counter =0
    for (key,Data_ROW)in Data_ROWS.items():
        if counter !=0:
            outputwb.create_sheet()
        ws = outputwb.worksheets[counter]
        ws.title = "社員番号"+str(key)
        fieldnames = ["ID","名前","開始日","終了日","申請種別","申請理由"]
        #ヘッダーの名前を登録する
        for iterator in range(0,len(fieldnames)):
            ws.cell(row=1, column=(iterator+1)).value = fieldnames[iterator]
    
        for iterator in range(0,len(Data_ROW)):
            Data = Data_ROW[iterator]
            ws.cell(row=(iterator+2), column=1).value = Data["ID"]
            ws.cell(row=(iterator+2), column=2).value = Data["NAME"]
            ws.cell(row=(iterator+2), column=3).value = Data["StartDate"]
            ws.cell(row=(iterator+2), column=4).value = Data["EndDate"]
            ws.cell(row=(iterator+2), column=5).value = Data["Title"]
            ws.cell(row=(iterator+2), column=6).value = Data["Reason"]
        counter+=1
    outputwb.save(OUTPUTFILENAME)
def ExcelWriter(parent,Data_ROW):
    ID = parent.St_ID.get()
    ToDay = datetime.datetime.now().strftime("%Y-%m-%d")
    OUTPUTFILENAME = os.getcwd()+"\OutPut\APPLYLIST_"+ID+"_"+ToDay+".xlsx"
    
    outputwb = excel.Workbook()
    ws = outputwb.active
    ws.title = "申請一覧"
    fieldnames = ["ID","名前","開始日","終了日","申請種別","申請理由"]
    #ヘッダーの名前を登録する
    for iterator in range(0,len(fieldnames)):
        ws.cell(row=1, column=(iterator+1)).value = fieldnames[iterator]
    
    for iterator in range(0,len(Data_ROW)):
        Data = Data_ROW[iterator]
        ws.cell(row=(iterator+2), column=1).value = Data["ID"]
        ws.cell(row=(iterator+2), column=2).value = Data["NAME"]
        ws.cell(row=(iterator+2), column=3).value = Data["StartDate"]
        ws.cell(row=(iterator+2), column=4).value = Data["EndDate"]
        ws.cell(row=(iterator+2), column=5).value = Data["Title"]
        ws.cell(row=(iterator+2), column=6).value = Data["Reason"]
    outputwb.save(OUTPUTFILENAME)


    
